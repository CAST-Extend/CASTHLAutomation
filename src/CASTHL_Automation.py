import shutil
import threading
import json
import datetime
import csv
import zipfile
from src import UnzipFile
from src import  AppRepoMapping
from src import HLScanAndOnboard
from pathlib import Path
import requests
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook
from src.logger_manager import LoggerManager
from src.AppRepoMapping import clean_folder_name
import os
import pandas as pd
import logging
import configparser
def get_all_repo_metadata(org_name, access_token, output_file_path, log_file_path):
    start_time = datetime.datetime.now()
    log_messages = []

    try:
        # Initialize an empty list to store all repositories
        all_repos = []

        # Fetch repositories from GitHub API with pagination
        headers = {
            "Authorization": f"Bearer {access_token}",
            "Accept": "application/vnd.github+json",
            "X-GitHub-Api-Version": "2022-11-28"
        }
        page_number = 1
        while True:
            repo_url = f"https://api.github.com/orgs/{org_name}/repos?per_page=200&page={page_number}"
            response = requests.get(repo_url, headers=headers)
            response.raise_for_status()  # Raise an exception for 4xx or 5xx status codes
            # print (response.raise_for_status)
            repos = response.json()
            if not repos:  # No more repositories on this page
                break
            all_repos.extend(repos)
            page_number += 1

        # Save repository metadata to JSON file
        with open(output_file_path, "w") as json_file:
            json.dump(all_repos, json_file, indent=4)

        log_messages.append(f"Metadata for all repositories in organization {org_name} downloaded successfully.")
        print(f"Metadata for all repositories in organization {org_name} downloaded successfully.")
    except requests.exceptions.RequestException as e:
        log_messages.append(f"Error: {str(e)}")
        if e.response.status_code==401:
            print('Bad credentials! Please check your "github_token" in config.properties file.')
        elif e.response.status_code==404:
            print('Bad Organization Name! Please check your "github_org_name" in config.properties file.')
        else:
            print(e)
        exit(0)

    end_time = datetime.datetime.now()
    total_time = end_time - start_time


    # Log the start and end time, along with total time taken
    with open(log_file_path, "a") as log_file:
        log_file.write(f"Start Time: {start_time}\n")
        log_file.write(f"End Time: {end_time}\n")
        log_file.write(f"Total Time Taken: {total_time}\n")
        for message in log_messages:
            log_file.write(message + "\n")

def json_to_csv(json_filename, csv_filename):
    try:
        # Read JSON data from file
        with open(json_filename, 'r') as json_file:
            json_data = json.load(json_file)
              
        # Add additional headers
        headers = ['id', 'name', 'default_branch', 'size', 'created_at', 'updated_at', 'pushed_at','clone_url','archive_url', 'batch_number']
        #headers.extend(additional_headers)
        
        # Write to CSV
        with open(csv_filename, 'w', newline='') as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=headers)
            writer.writeheader()
            count = 0
            batch_num = 1
            for entry in json_data:
                # Extract data from each entry
                row_data = {key: entry.get(key, '') for key in headers}

                count=count+1
                if count % 500 == 0:
                    batch_num = batch_num + 1
                row_data['batch_number'] = batch_num
                # print(row_data)

                # Write row to CSV
                writer.writerow(row_data)
        print("Repo Summary CSV file generated successfully.")
        
    except FileNotFoundError:
        print("File not found.")
    except json.JSONDecodeError:
        print("Invalid JSON format.")
    except Exception as e:
        print(f"An error occurred: {e}")

def modify_archive_urls(csv_file_path):
    try:
        # Read the CSV file
        df = pd.read_csv(csv_file_path, encoding='latin-1')

        # Define a function to replace placeholders in archive_url and create a new column
        def replace_url(row):
            archive_format = 'zipball/'
            ref = row['default_branch']
            final_url = row['archive_url'].replace('{archive_format}', archive_format).replace('{/ref}', ref)
            return final_url

        # Apply the function to create the new column 'archive_url_download'
        df['repo_archive_download_api'] = df.apply(replace_url, axis=1)

        # Write the modified DataFrame back to the original CSV file
        df.to_csv(csv_file_path, index=False)
    except Exception as e:
        print(f"Error while executing modify_archive_urls() function: {e}")

def add_new_columns_to_csv(output_csv_file_path):
    try:
        df = pd.read_csv(output_csv_file_path, encoding='latin-1')

        # Add two new columns with some default values or calculations
        df['Download'] = 'N'  # You can also use a calculation or other values
        df['Download_Status'] = ''

        # Path to save the updated CSV file
        df.to_csv(output_csv_file_path, index=False)
    except Exception as e:
        print(f"Error while executing add_new_columns_to_csv() function: {e}")   
def update_rescan_column(mapping_excel_path,output_csv_file_path,logger):
    try:
        logger.info("Starting update_rescan_column()...")
        logger.info(f"Reading source CSV: {output_csv_file_path}")
        source_df = pd.read_csv(output_csv_file_path)

        logger.info(f"Reading target Excel: {mapping_excel_path}")
        target_df = pd.read_excel(mapping_excel_path)

        if 'name' not in source_df.columns or 'Download' not in source_df.columns:
            logger.error("Missing required columns ['name', 'Download'] in source CSV.")
            return

        if 'Repository' not in target_df.columns or 'Application' not in target_df.columns:
            logger.error("Missing required columns ['Repository', 'Application'] in target Excel.")
            return

        logger.info("Creating mapping between source and target...")
        download_map = dict(zip(source_df['name'], source_df['Download']))

        target_df['rescan'] = target_df['Repository'].map(download_map)
        target_df['rescan'] = target_df['rescan'].fillna('N')

        logger.info("Saving updated Excel file...")
        target_df.to_excel(mapping_excel_path, index=False)

        rescan_apps = target_df.loc[target_df['rescan'] == 'Y', 'Application'].unique().tolist()
        logger.info(f"Applications that need rescan ({len(rescan_apps)} found): {rescan_apps}")
    except Exception as e:
        logger.exception(f"Error while executing update_rescan_column(): {e}")
def update_download_column(output_csv_file_path, last_refresh_date):
    try:
        # Load your CSV file
        df = pd.read_csv(output_csv_file_path, encoding='latin-1')

        # Convert pushed_at to datetime (handles ISO format with Z)
        df['pushed_at'] = pd.to_datetime(df['pushed_at'], utc=True)

        # Clean and parse last_refresh_date string
        last_refresh_date = datetime.datetime.strptime(last_refresh_date.strip('"'), '%Y-%m-%d').replace(tzinfo=datetime.timezone.utc)

        # Compare and update 'Download' column
        df['Download'] = df['pushed_at'].apply(lambda x: 'Y' if x > last_refresh_date else 'N')

        # Save updated CSV
        df.to_csv(output_csv_file_path, index=False)

        print("Download column updated based on pushed_at date.")
    except Exception as e:
        print(f"Error while executing update_download_column(): {e}")

def add_action_column(mapping_excel_path, repo_summary_csv_path):
    try:
        # Load Excel and CSV files
        df_mapping = pd.read_excel(mapping_excel_path)
        df_summary = pd.read_csv(repo_summary_csv_path)

        # Normalize values for matching
        df_mapping['Repository'] = df_mapping['Repository'].str.strip()
        df_summary['name'] = df_summary['name'].str.strip()

        # Build dictionary for lookup: repo name -> Download status
        repo_download_map = df_summary.set_index('name')['Download'].to_dict()

        # Determine Action
        def get_action(repo):
            if repo in repo_download_map:
                return 'Replaced' if repo_download_map[repo] == 'Y' else 'Skipped'
            return 'Deleted'

        # Apply and update Action column
        df_mapping['Action'] = df_mapping['Repository'].apply(get_action)

        # Overwrite the same Excel file
        df_mapping.to_excel(mapping_excel_path, index=False)
        print(f"Updated Action column in Excel file: {mapping_excel_path}\n")

    except Exception as e:
        print(f"Error while executing add_action_column() function: {e}")   

def check_column_exists(file_path, column_name):
    try:
        # Read the CSV file
        df = pd.read_csv(file_path, encoding='latin-1')
        
        # Check if the column exists
        if column_name not in df.columns:
            print(f"Column '{column_name}' does not exist in the CSV file.")
            return False
        
        # Check if the column has data
        if df[column_name].empty:
            print(f"Column '{column_name}' exists but has no data.")
            return False
        
        return True
    except FileNotFoundError:
        print(f"File '{file_path}' not found.")
        return False
    except Exception as e:
        print(f"An error occurred: {e}")
        return False
    
def read_csv_data(file_path):
    """
    Reads data from a CSV file.
    Parameters:
        file_path (str): The path to the CSV file.
    Returns:
        list: A list containing tuples of data extracted from the CSV file.
    """
    data = []
    try:
        with open(file_path, mode='r', newline='', encoding='latin-1') as file:
            reader = csv.reader(file)
            next(reader)  # Skip the header row
            for row in reader:
                data.append((row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11], row[12]))  # Assuming 4 columns in the CSV
    except Exception as e:
        print(f"Error executing read_csv_data() function: {e}")
    return data

def log_start_end_time(repository_name, start_time, end_time, total_time, log_file):
    """
    Logs start and end time of a process.
    Parameters:
        repository_name (str): The name of the repository or process.
        start_time (datetime): The start time of the process.
        end_time (datetime): The end time of the process.
        total_time (timedelta): The total time taken for the process.
        log_file (str): The path to the log file.
    """
    try:
        log_message = f"{repository_name} | {start_time} | {end_time} | {total_time} |"
        with open(log_file, "a") as f:
            f.write(log_message + "\n")
    except Exception as e:
        print(f"Error while executing function log_start_end_time():", {e})

def log_processing(repository_name, status, log_file):
    """
    Logs the processing status of a repository.
    Parameters:
        repository_name (str): The name of the repository.
        status (str): The processing status.
        log_file (str): The path to the log file.
    """
    try:    
        log_message = f"{repository_name} | {status}"
        with open(log_file, "a") as f:
            f.write(log_message + "\n")
    except Exception as e:
        print(f"Error while executing log_processing() function: {e}") 

def download_zip_archive(repository_url, repository_path, token):
    """
    Downloads a ZIP archive from a given URL.
    Parameters:
        repository_url (str): The URL of the repository.
        repository_path (str): The path to save the ZIP archive.
        token (str): The GitHub access token.
        
    Returns:
        bool: True if download is successful, False otherwise.
    """
    #print(f"Inside **download_zip_archive**'.")
    try:
        headers = {'Authorization': f'token {token}'}
        response = requests.get(repository_url, headers=headers)
        reason = None
        
        if response.status_code == 200:
            with open(repository_path, 'wb') as f:
                f.write(response.content)
            return True, reason
        else:
            reason = str(response.status_code) +' '+ str(response.reason)
            return False, reason
    except Exception as e:
        print(f"Error while executing download_zip_archive() function: {e}")
  
def update_download_status(repo_id, download_status, output_csv_file_path_batch):
    try:
        # Read the CSV file
        with open(output_csv_file_path_batch, 'r', newline='', encoding='latin-1') as file:
            reader = csv.DictReader(file)
            rows = list(reader)
        
        # Modify the data in the specified column
        for row in rows:
            if row['id']==repo_id:
                row['Download_Status'] = download_status
        
        # Write the updated data back to the CSV file
        with open(output_csv_file_path_batch, 'w', newline='', encoding='latin-1') as file:
            writer = csv.DictWriter(file, fieldnames=reader.fieldnames)
            writer.writeheader()
            writer.writerows(rows)
    except Exception as e:
        print(f"Error while executing update_download_status() function: {e}")

def split_csv_based_on_batch_num(batch_num, output_csv_file_path, output_csv_file_path_batch):
    try:

        df = pd.read_csv(output_csv_file_path)

        batch_df = df[df['batch_number'] == batch_num]
        batch_df.to_csv(output_csv_file_path_batch, index=False)

        print(f"{output_csv_file_path_batch} file created:")

    except Exception as e:
        print(f"Error while executing split_csv_based_on_batch_num() function: {e}")

def combine_all_batch_csv_files(output_csv_file_path_batch, output_csv_file_path):
    try:
        # List all CSV files in the input directory
        csv_files = [os.path.join(output_csv_file_path_batch, file) for file in os.listdir(output_csv_file_path_batch) if file.endswith('.csv')]
        # Combine all CSV files into one DataFrame
        combined_df = pd.concat([pd.read_csv(f) for f in csv_files])
        # Save the combined DataFrame to a CSV file
        combined_df.to_csv(output_csv_file_path, index=False)
        print(f"Combined CSV saved to {output_csv_file_path}")

    except Exception as e:
        print(f"Error while executing combine_all_batch_csv_files() function: {e}")


def download_and_save_code(application_name, download_status, repository_url, server_location, token, start_end_log_file, processing_log_file, output_csv_file_path, repo_id, output_csv_file_path_batch):
    """
    Downloads and saves code from a repository.
    Parameters:
        application_name (str): The name of the application or repository.
        repository_url (str): The URL of the repository.
        server_location (str): The location to save the repository.
        token (str): The GitHub access token.
        start_end_log_file (str): The path to the log file for start and end times.
        processing_log_file (str): The path to the log file for processing status.
    """
    try:
        #print(f"Inside **Download-And-Save**'.")
        application_name_directory = os.path.join(server_location, application_name)
            # Check if the 'Output' folder exists, if not, create it
        if not os.path.exists(application_name_directory):
            os.makedirs(application_name_directory)
        else:
            dir_to_delete = application_name_directory
            command = f'rmdir /s /q "{dir_to_delete}"'
            os.system(command)
            os.makedirs(application_name_directory)
        
        repository_zip_path = os.path.join(application_name_directory, application_name + '.zip')
        #print(f"repository_zip_path '{repository_zip_path}'.")
        if os.path.exists(repository_zip_path):
            log_processing(application_name, "Skipped: ZIP file already exists", processing_log_file)
            print(f"Skipping repository '{application_name}'. ZIP file already exists.\n")
            
        else:
            start_time = datetime.datetime.now()
            try:
                download_flag, reason = download_zip_archive(repository_url, repository_zip_path, token)
                if download_flag:
                    with zipfile.ZipFile(repository_zip_path, 'r') as zip_ref:
                        file_list = zip_ref.namelist()
                        if not file_list:
                            log_processing(application_name, "Repo is empty", processing_log_file)
                            print(f"Repository '{application_name}' is empty.\n")
                        else:
                            end_time = datetime.datetime.now()
                            total_time = end_time - start_time
                            log_start_end_time(application_name, start_time, end_time, total_time, start_end_log_file)
                            log_processing(application_name, "Successful", processing_log_file)
                            print(f"Repository '{application_name}' downloaded successfully as ZIP file to '{repository_zip_path}'.\n")
                            download_status = 'Success'
                            update_download_status(repo_id, download_status, output_csv_file_path_batch)

                else:
                    end_time = datetime.datetime.now()
                    total_time = end_time - start_time
                    log_start_end_time(application_name, start_time, end_time, total_time, start_end_log_file)
                    log_processing(application_name, "Failed - "+reason, processing_log_file)
                    print(f"Failed to download repository '{application_name}', Because of the reason - {reason}.\n")
                    download_status = f'Failed - {reason}'
                    update_download_status(repo_id, download_status, output_csv_file_path_batch)

            except Exception as e:
                end_time = datetime.datetime.now()
                total_time = end_time - start_time
                log_start_end_time(application_name, start_time, end_time, total_time, start_end_log_file)
                log_processing(application_name, f"Failed: {e}", processing_log_file)
                print(f"Error downloading repository: {e}")
    except Exception as e:
        print(f"Error while executing download_and_save_code() function: {e}")

def download_in_batch(batch, thread_id, src_dir, token, start_end_log_file, processing_log_file, output_csv_file_path, output_csv_file_path_batch):
    try:
        # thread_log_file = f"Repos_download_thread_{thread_id}_{datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log"
        # logging.basicConfig(filename=thread_log_file, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        # logging.info(f'Thread {thread_id} started.\n')
        print(f'Thread {thread_id} started.\n')

        start_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # logging.info(f'Thread {thread_id} start time: {start_time}\n')
        print(f'Thread {thread_id} start time: {start_time}\n')
        # logging.info(f'Thread {thread_id} processing repos: {batch}\n')
        # print(f'Thread {thread_id} processing repos: {batch}')

        for repository in batch:
            if repository[11] == 'Y':
                download_and_save_code(repository[1], repository[12], repository[10], src_dir, token, start_end_log_file, processing_log_file, output_csv_file_path, repository[0], output_csv_file_path_batch)
            else:
                print(f"User Marked Download='N' Hence Skipping the Download of Repo -> '{repository[1]}'\n")

        end_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # logging.info(f'Thread {thread_id} end time: {end_time}\n')
        print(f'Thread {thread_id} end time: {end_time}\n')
        # logging.info(f'Thread {thread_id} finished.\n')
        print(f'Thread {thread_id} finished.\n')
    except Exception as e:
        print(f"Error while executing download_in_batch() function: {e}")

def mainframeCopyAppend_to_analyzed_from_csv(csv_file_path, log_file_path):
    log_messages = []

    try:
        df = pd.read_csv(csv_file_path)

        if 'Mainframe_src_folder' not in df.columns or 'Destination_folder' not in df.columns:
            raise ValueError("CSV must contain 'Mainframe_src_folder' and 'Destination_folder' columns.")

        for index, row in df.iterrows():
            src_folder = str(row['Mainframe_src_folder']).strip()
            dest_folder = str(row['Destination_folder']).strip()

            folder_name = os.path.basename(src_folder.rstrip("\\/"))
            final_dest_path = os.path.join(dest_folder, folder_name)

            log_messages.append(f"{datetime.datetime.now()} - INFO - Copying: {src_folder} --> {final_dest_path}")

            if not os.path.exists(src_folder):
                log_messages.append(f"{datetime.datetime.now()} - ERROR - Source folder does not exist: {src_folder}")
                continue

            try:
                # Create destination parent if not exists
                os.makedirs(dest_folder, exist_ok=True)
                # Copy the folder itself inside destination
                shutil.copytree(src_folder, final_dest_path, dirs_exist_ok=True)
                log_messages.append(f"{datetime.datetime.now()} - SUCCESS - Copied {src_folder} to {final_dest_path}")
            except Exception as copy_error:
                log_messages.append(f"{datetime.datetime.now()} - ERROR - Failed to copy folder: {copy_error}")

    except Exception as e:
        log_messages.append(f"{datetime.datetime.now()} - CRITICAL - Failed to process CSV: {e}")
 
    # Log messages to the specified log file
    with open(log_file_path, "a") as log_file:
        for message in log_messages:
            log_file.write(message + "\n")

def find_long_paths(parent_folder, max_length=260):
    try:
        long_paths = []

        applications = []
        for app_name in os.listdir(parent_folder):
            if os.path.isdir(os.path.join(parent_folder, app_name)):
                app_path = os.path.join(parent_folder, app_name)
                applications.append((app_name, app_path))
        # print(applications)

        for app_name, app_path in applications:
            # Walk through the directory tree
            for root, dirs, files in os.walk(app_path):
                for file in files:
                    # Construct the full path of the file
                    file_path = os.path.join(root, file)
                    # Check if the path length exceeds the limit
                    if len(file_path) > max_length:
                        long_paths.append((app_name,file_path))
        return long_paths
    except Exception as e:
        print(f"Error while executing find_long_paths() function: {e}")

def create_app_txt(latest_hl_data, rescan_summary, app_logger, txt_output_path):
    """
    Compare Troux UUID from rescan summary with Application ClientRef from Highlight data.
    If matched, write Application Name and Application ID to a text file.

    Output format:
    Application Name;Application ID
    """
    try:
        app_logger.info("=== Step: Create App List TXT Started ===")
        app_logger.info(f"Reading Highlight Excel: {latest_hl_data}")
        app_logger.info(f"Reading Rescan Summary Excel: {rescan_summary}")

        # Read Excel files
        hl_df = pd.read_excel(latest_hl_data, dtype=str)
        rescan_df = pd.read_excel(rescan_summary, dtype=str)

        # Clean whitespace
        hl_df["Application ClientRef"] = hl_df["Application ClientRef"].astype(str).str.strip()
        rescan_df["Troux UUID"] = rescan_df["Troux UUID"].astype(str).str.strip()

        # Merge data
        merged_df = pd.merge(
            rescan_df,
            hl_df,
            left_on="Troux UUID",
            right_on="Application ClientRef",
            how="inner"
        )

        app_logger.info(f"Matched applications found: {len(merged_df)}")

        # Prepare the output lines with header
        output_lines = ["Application Name;Application ID"]

        # Add application data only if matches exist
        if not merged_df.empty:
            output_lines.extend([
                f"{row['Application Name']};{row['Application ID']}"
                for _, row in merged_df.iterrows()
            ])
        else:
            app_logger.info("No matching applications found. Only header will be written.")

        # Ensure output directory exists
        os.makedirs(os.path.dirname(txt_output_path), exist_ok=True)

        # Write to txt file
        with open(txt_output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(output_lines))

        app_logger.info(f"Application list written to: {txt_output_path}")
        app_logger.info("=== Step Completed Successfully ===")

        print(f"\nOutput TXT generated at: {txt_output_path}")

    except Exception as e:
        app_logger.error(f"Error while creating app txt: {e}", exc_info=True)
        print(f"Error while creating app txt: {e}")


def fetch_and_save_applications(url, headers, output_path, logger,current_datetime):
    """
    Fetch applications from a REST API and save them to an Excel file
    with formatted headers and auto-adjusted column widths.

    Parameters:
        url (str): API endpoint for fetching applications
        headers (dict): Authorization or other HTTP headers
        output_path (str): Path where the Excel file will be saved
        log_file_path (str): Optional file path for logging output
    """

    try:
        logger.info("Fetching applications from HL Instance...")
        response = requests.get(url, headers=headers)
        response.raise_for_status()

        apps = response.json()
        if not apps:
            logger.warning("No applications found in response.")
            return

        data = []
        for app in apps:
            data.append({
                "Application ID": app.get("id"),
                "Application Name": app.get("name"),
                "Application ClientRef": app.get("clientRef", None)
            })

        df = pd.DataFrame(data)
        output_path = os.path.join(output_path, f"LM_Highlight_AppDetails_{current_datetime}.xlsx")
        logger.info("Found {} Applications in the HL instance".format(len(data)))
        # Save to Excel
        df.to_excel(output_path, index=False)
        logger.info(f"Applications Data written to Excel: {output_path}")

        # Load workbook for styling
        wb = load_workbook(output_path)
        ws = wb.active

        # Header styling
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")  # Dark blue
        header_font = Font(color="FFFFFF", bold=True)  # White bold text
        center_align = Alignment(horizontal="center", vertical="center")

        # Apply style to header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        # Auto column width
        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 4

        wb.save(output_path)
        logger.info(f"Applications list  saved successfully: {output_path}")

    except requests.exceptions.RequestException as e:
        logger.exception(f"API request failed: {e}")
    except Exception as e:
        logger.exception(f"Error while creating Excel: {e}")



def match_applications(metadata_file, mapping_file, output_file, logger=None):
    """
    Match applications from metadata and mapping files where Download == 'Y',
    and export unique Application–Troux UUID pairs to an Excel file.

    Parameters:
        metadata_file (str): Path to the metadata CSV file
        mapping_file (str): Path to the App-Repo-Mapping Excel file
        output_file (str): Path where matched output Excel will be saved
        logger (logging.Logger, optional): Logger for logging info/warnings
    """

    try:
        if logger:
            logger.info("Starting application matching process...")
            logger.info(f"Metadata file: {metadata_file}")
            logger.info(f"Mapping file: {mapping_file}")
        else:
            print("Starting application matching process...")

        # Step 1: Load input files
        metadata_df = pd.read_csv(metadata_file)
        mapping_df = pd.read_excel(mapping_file)
        if logger:
            logger.info(f"Loaded {len(metadata_df)} metadata records and {len(mapping_df)} mapping records.")

        # Step 2: Filter rows where Download == 'Y'
        metadata_filtered = metadata_df[metadata_df['Download'].str.upper() == 'Y']
        if logger:
            logger.info(f"Filtered {len(metadata_filtered)} records with Download='Y'.")

        # Step 3: Merge like VLOOKUP (name ↔ Repo Name)
        merged_df = pd.merge(
            metadata_filtered,
            mapping_df,
            how='inner',
            left_on='name',
            right_on='Repository'
        )
        if logger:
            logger.info(f"Merged records: {len(merged_df)}")

        # Step 4: Select required columns
        output_df = merged_df[['Application', 'Troux UUID']].copy()

        # Step 5: Remove duplicates
        before_dedup = len(output_df)
        output_df = output_df.drop_duplicates(subset=['Application', 'Troux UUID'], keep='first')
        after_dedup = len(output_df)
        if logger:
            logger.info(f"Removed {before_dedup - after_dedup} duplicate entries.")

        output_df['Application'] = output_df['Application'].apply(
            lambda x: clean_folder_name(str(x).strip())
        )
        # Step 7: Export to Excel
        output_df.to_excel(output_file, index=False)

        msg = f"Output saved to: {output_file} | Total unique matched records: {len(output_df)}"
        if logger:
            logger.info(msg)
        else:
            print(msg)

    except Exception as e:
        error_msg = f"Error in match_applications(): {e}"
        if logger:
            logger.exception(error_msg)
        else:
            print(error_msg)



def create_applications_hl(hl_url, logger, App_Repo_Mapping, token,highlight_company_id):
    """
    Read 'NewApplications' sheet from App_Repo_Mapping Excel and create applications in CAST Highlight.

    Parameters:
        hl_url (str): Base URL of CAST Highlight instance
        logger: Logger instance
        App_Repo_Mapping (str): Path to Excel file containing 'NewApplications' sheet
        token (str): Bearer token for authorization
    """
    try:
        # Step 1: Read NewApplications sheet
        df_new_apps = pd.read_excel(App_Repo_Mapping, sheet_name='NewApplications')
        logger.info(f"Found {len(df_new_apps)} new applications from New Applications List To creatwe in HL instance.")

        # Step 2: Loop through each row and create application
        for index, row in df_new_apps.iterrows():
            app_name = str(row['Application'])
            clientref = str(row['Troux UUID'])

            url = f"{hl_url}/WS2/portfolioManagement/domains/{highlight_company_id}/applications?expand=contributor&expand=domain"

            payload ={
          "name": app_name,
          "contributors": [],
          "domains": [
            {
              "id": highlight_company_id
            }
          ],
          "status": "notArchived",
          "clientRef": clientref,
          "componentIds": []
        }

            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {token}"
            }

            response = requests.post(url, json=payload, headers=headers)

            if response.status_code in (200, 201):
                msg = f"[{index + 1}] Application '{app_name}' created successfully in the Hl."
                logger.info(msg)
            else:
                msg = f"[{index + 1}] Failed to create application '{app_name}' in the Hl. Status: {response.status_code}, Response: {response.text}"
                logger.error(msg)

    except Exception as e:
        error_msg = f"Error in create_applications_hl_from_excel(): {e}"
        if logger:
            logger.exception(error_msg)



def find_new_applications(app_list_file, mapping_file, logger=None):
    try:
        if logger:
            logger.info("Starting applications comparison to find new appliactiosn in App-Repo-Mapping")
            logger.info(f"Application list file: {app_list_file}")
            logger.info(f"Mapping file: {mapping_file}")

        # Step 1: Read both Excel files
        app_list_df = pd.read_excel(app_list_file)
        mapping_df = pd.read_excel(mapping_file)

        # Step 2: Validate required columns
        required_cols_app = {'Application Name', 'Application ClientRef'}
        required_cols_map = {'Application', 'Troux UUID'}

        if not required_cols_app.issubset(app_list_df.columns):
            raise ValueError(f"Application List file must contain columns: {required_cols_app}")
        if not required_cols_map.issubset(mapping_df.columns):
            raise ValueError(f"Mapping file must contain columns: {required_cols_map}")

        # Step 3: Normalize data (lowercase + trim)
        app_list_df['Application Name'] = app_list_df['Application Name'].astype(str).str.strip()
        app_list_df['Application ClientRef'] = app_list_df['Application ClientRef'].astype(str).str.strip()

        mapping_df['Application'] = mapping_df['Application'].astype(str).str.strip()
        mapping_df['Troux UUID'] = mapping_df['Troux UUID'].astype(str).str.strip()

        # Step 4: Compare mapping data with application list
        merged_df = pd.merge(
            mapping_df,
            app_list_df,
            how='left',
            left_on=['Application', 'Troux UUID'],
            right_on=['Application Name', 'Application ClientRef'],
            indicator=True
        )

        # Step 5: Identify new applications (not found in app list)
        new_apps_df = merged_df[merged_df['_merge'] == 'left_only'][['Application', 'Troux UUID']].copy()
        new_apps_df['Application'] = new_apps_df['Application'].apply(
            lambda x: clean_folder_name(str(x).strip())
        )

        if not new_apps_df.empty:
            with pd.ExcelWriter(mapping_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                new_apps_df.to_excel(writer, sheet_name='NewApplications', index=False)

            msg = f"Found {len(new_apps_df)} new/unmapped applications. Stored in 'NewApplications' sheet of {mapping_file}"
        else:
            msg = "No new/unmapped applications found."

        if logger:
            logger.info(msg)

    except Exception as e:
        error_msg = f"Error in find_new_applications(): {e}"
        if logger:
            logger.exception(error_msg)


def get_hl_applications_path(config_dir):
    for file_name in os.listdir(config_dir):
        if file_name.startswith("LM_Highlight_AppDetails_") and file_name.endswith(".xlsx"):
            return os.path.join(config_dir, file_name)
    return None

def identify_to_be_deleted_apps(hl_applications_path,output_dir,logs_dir,App_Repo_Mapping):
    """
    Identify Highlight applications that no longer exist in GitHub (candidate for deletion).

    Steps:
      1. Compare 'CLIENTREF' in LM_Highlight_AppDetails_*.xlsx (Applications sheet)
         vs 'Troux UUID' in App2RepoMapping.xlsx.
      2. Generate Excel output with only [CLIENTREF, Name].
      3. Log progress and results.

    Returns:
        output_file (str): Path to the generated Excel file.
        log_file (str): Path to the log file.
    """
    from datetime import datetime


    def setup_logger(log_dir):
        """Setup logger with timestamp"""
        os.makedirs(log_dir, exist_ok=True)
        log_file = os.path.join(
            log_dir,
            f"Identify_ToBeDeleted_Apps_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.log")

        logging.basicConfig(
            filename=log_file,
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s"
        )
        return log_file


    os.makedirs(output_dir, exist_ok=True)
    log_file = setup_logger(logs_dir)

    try:
        logging.info("=== Step: Identify To-Be-Deleted Applications Started ===")

        # Validate input files
        if not os.path.exists(hl_applications_path):
            raise FileNotFoundError(f"Highlight file not found: {hl_applications_path}")
        if not os.path.exists(hl_applications_path):
            raise FileNotFoundError(f"Mapping file not found: {hl_applications_path}")

        # Load Highlight Excel
        logging.info(f"Loading Highlight data from: {hl_applications_path}")
        hl_df = pd.read_excel(hl_applications_path, dtype=str)
        hl_df["Application ClientRef"] = hl_df["Application ClientRef"].astype(str).str.strip()

        # Load Mapping Excel
        logging.info(f"Loading mapping data from: {hl_applications_path}")
        map_df = pd.read_excel(App_Repo_Mapping, dtype=str)
        map_df["Troux UUID"] = map_df["Troux UUID"].astype(str).str.strip()

        # Compare and filter missing apps
        missing_apps = hl_df[~hl_df["Application ClientRef"].isin(map_df["Troux UUID"])]
        output_df = missing_apps[["Application ClientRef", "Application Name"]].copy()
        output_df.rename(columns={"Application ClientRef": "TrouxID", "Name": "Application Name"}, inplace=True)
        # Output summary
        logging.info(f"Total Highlight apps: {len(hl_df)}")
        logging.info(f"Total Mapping apps: {len(map_df)}")
        logging.info(f"Candidate apps for deletion: {len(output_df)}")

        # Output file
        month_str = datetime.now().strftime("%b")
        output_file = os.path.join(output_dir, f"LM_HL_Refresh_{month_str}_TobeDeleted_Apps.xlsx")

        # Write result
        output_df.to_excel(output_file, index=False)
        logging.info(f"Output written to: {output_file}")
        logging.info("=== Step Completed Successfully ===")

        print(f"\nOutput generated: {output_file}")
        print(f"Log file: {log_file}")
        return output_file, log_file

    except Exception as e:
        logging.exception("Error occurred during Identify_ToBeDeleted_Apps process")
        print(f" Error: {e}")
        return None, log_file
def main():
    while True:
        current_datetime = datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

        # Read configuration from config.properties file
        config = configparser.ConfigParser()
        config_file_path = os.path.join(os.path.dirname(__file__), '..', 'Config', 'config.properties')
        config.read(config_file_path)

        # Get values from the config file
        org_name = config.get('GitHub', 'github_org_name')
        token = config.get('GitHub', 'github_token')
        config_dir = config.get('Directories', 'config_dir')
        src_dir = config.get('Directories', 'src_dir')
        unzip_dir = config.get('Directories', 'unzip_dir')
        logs_dir = config.get('Directories', 'logs_dir')
        output_dir = config.get('Directories', 'output_dir')
        App_Repo_Mapping = config.get('Input-File', 'App_Repo_Mapping')
        #mainframe_src_folder = config.get('Directories', 'mainframe_src_folder')
        src_dir_analyze = config.get('Directories', 'src_dir_analyze')
        last_refresh_date = config.get('GitHub', 'last_refresh_date')
        csv_file_path = config.get('Input-File', 'source_destination_mapping')  # Read the CSV file path
        highlight_base_url=config.get('HIGHLIGHT-ONBOARDING','highlight_base_url')
        highlight_company_id=config.get('HIGHLIGHT-ONBOARDING','highlight_company_id')
        highlight_token=config.get('HIGHLIGHT-ONBOARDING','highlight_token')
        highlight_application_mapping=config.get('HIGHLIGHT-ONBOARDING','highlight_application_mapping')
        # Check if the 'Source Dir' folder exists, if not, create it
        if not os.path.exists(src_dir):
            os.makedirs(src_dir)

        # Check if the 'unzip_dir' folder exists, if not, create it
        if not os.path.exists(unzip_dir):
            os.makedirs(unzip_dir)
        
        # Check if the 'Log Dir' folder exists, if not, create it
        if not os.path.exists(logs_dir):
            os.makedirs(logs_dir)
        
        # Check if the 'Output' folder exists, if not, create it
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Check if the 'Output' folder exists, if not, create it
        if not os.path.exists(src_dir_analyze):
            os.makedirs(src_dir_analyze)

        while True:
            print("Select options:")
            print("1. Download Metadata for GitHub organization")
            print("2. List Out Candidate Applications For Rescan")
            print("3. Export HL existing data")
            print("4. Identify and create new applications")
            print("5. Identify Deleted Applications")
            print("6. Download source code for all repositories in the organization in batches")
            print("7. Unzip the downloaded source code")
            print("8. Create application folders and move repositories")
            print("9. Copy Mainframe folder from src to dest")
            print("10. Prepare Application.txtt")
            print("11. Trigger CAST Highlight onboarding for the source code")
            choice = input("Enter your choice (1/2/3/4/5/6/7/8/9/10/11): ")

            if choice not in ['1', '2', '3', '4', '5', '6', '7', '8','9','10','11']:
                print("Invalid choice. Please enter 1, 2, 3, 4, 5, 6, 7, 8,9,10,11")
                continue

            output_type = int(choice)
            main_operations(output_type, current_datetime, org_name, token, config_dir, src_dir, unzip_dir, logs_dir, output_dir, App_Repo_Mapping, csv_file_path, src_dir_analyze, last_refresh_date,highlight_base_url,highlight_company_id,highlight_token,highlight_application_mapping)

            # Ask user if they want to continue
            continue_option = input("Do you want to run another query? (yes/no): ")
            if continue_option.lower() != 'yes':
                exit(0)

def main_operations(output_type, current_datetime, org_name, token, config_dir, src_dir, unzip_dir, logs_dir, output_dir, App_Repo_Mapping, csv_file_path, src_dir_analyze, last_refresh_date,highlight_base_url,highlight_company_id,highlight_token,highlight_application_mapping):
    if output_type == 1:
        output_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Metadata.json")
        log_file_path = os.path.join(logs_dir, f"{org_name}_Metadatadownload_{current_datetime}.log")
        output_csv_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Summary.csv")
        get_all_repo_metadata(org_name, token, output_file_path, log_file_path)
        json_to_csv(output_file_path, output_csv_file_path)
        modify_archive_urls(output_csv_file_path)
        add_new_columns_to_csv(output_csv_file_path)
        update_download_column(output_csv_file_path, last_refresh_date)
        print(f"Refer Log file {log_file_path} for download log and time to download Metadata.")
        print(f"CSV file generated {output_csv_file_path} with summary of repositories which can be used for downloading source code(Task-2).\n")

    elif output_type == 6:
        output_csv_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Summary.csv")
        if not os.path.exists(output_csv_file_path):
            print("Please run option 1 to download metadata first.")
            return
        if not check_column_exists(output_csv_file_path, 'batch_number'):
            print(f"Column 'batch_number' does not exist in file {output_csv_file_path}. Create a new column with the name 'batch_number' and enter the number of the batch you want to run for download.")
            return
        if not check_column_exists(output_csv_file_path, 'repo_archive_download_api'):
            print(f"Column 'repo_archive_download_api' does not exist in file {output_csv_file_path}. Review the CSV file and rerun option 1.")
            return
        if not check_column_exists(output_csv_file_path, 'name'):
            print(f"Column 'name' does not exist in file {output_csv_file_path}. Review the CSV file and rerun option 1.")
            return
        output_csv_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Summary.csv")
        data = read_csv_data(output_csv_file_path) 
        # print(data[0])
        batches = []
        num_of_batches = int(data[-1][9])
        for i in range(num_of_batches):
            batch = []
            for repository in data:
                if int(repository[9]) == i+1:
                    batch.append(repository)
            batches.append(batch)
        # Process batches using multi-threading
        threads = []
        # Define the directory name
        directory = Path(output_dir+"\csv_for_each_batch")
        # Create the directory
        directory.mkdir(parents=True, exist_ok=True)
        print(f"Directory '{directory}' created successfully.")

        if os.path.exists(directory):
            # List all files in the directory
            files_in_directory = os.listdir(directory)
            # Loop through the files and delete each one
            for file in files_in_directory:
                file_path = os.path.join(directory, file)
                # Check if it's a file and delete it
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    print(f"Deleted file: {file_path}")
        
        for i, batch in enumerate(batches, start=1):
            start_end_log_file = os.path.join(logs_dir, f"RepoDownloadTime_batch_{i}_{current_datetime}.txt")
            processing_log_file = os.path.join(logs_dir, f"RepoDownloadStatusLog_batch_{i}_{current_datetime}.txt")
            output_csv_file_path_batch = os.path.join(directory, f"{org_name}_Repositories_Summary_batch_{i}.csv")
            split_csv_based_on_batch_num(i, output_csv_file_path, output_csv_file_path_batch)

            if not os.path.exists(start_end_log_file):
                with open(start_end_log_file, "w") as start_end_log:
                    start_end_log.write("Start Time\tEnd Time\tTotal Time Taken\n")            
            if not os.path.exists(processing_log_file):
                with open(processing_log_file, "w") as processing_log:
                    processing_log.write("Timestamp\tMessage\n")
            open(start_end_log_file, 'w').close()
            open(processing_log_file, 'w').close()
            thread = threading.Thread(target=download_in_batch, args=(batch, i, src_dir, token, start_end_log_file, processing_log_file, output_csv_file_path, output_csv_file_path_batch))
            threads.append(thread)
        # Start threads
        for t in threads:
            t.start()
        # Join threads to the main thread
        for t in threads:
            t.join()

        combine_all_batch_csv_files(directory, output_csv_file_path)       
    elif output_type==2:
        rescan_logger = LoggerManager.get_logger("Applications_rescan", log_dir=logs_dir)
        output_csv_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Summary.csv")
        output_file=os.path.join(output_dir, f"Rescan_applications_summary.xlsx")
        match_applications(output_csv_file_path,App_Repo_Mapping,output_file,rescan_logger)
    elif output_type == 7:
        try:
            UnzipFile.unzip_code(src_dir, unzip_dir, os.path.join(logs_dir, f"Unzip_Execution_{current_datetime}.log"), os.path.join(logs_dir, f"Unzip_Time_{current_datetime}.log"))
        except Exception as e:
            print(f"Error occurred during extraction: {e}")
    elif output_type==3:
        rescan_logger = LoggerManager.get_logger("Export_HL_existing_data", log_dir=logs_dir)
        url = "{}/WS2/domains/{}/applications".format(highlight_base_url,highlight_company_id)
        headers = {"Authorization": "Bearer {}".format(highlight_token)}
        from datetime import datetime
        current_date = datetime.now().strftime("%Y%m%d")
        fetch_and_save_applications(url,headers,output_dir,rescan_logger,current_date)
    elif output_type == 8:
        log_file=os.path.join(logs_dir, f"AppRepoMapping_{current_datetime}.log")
        logger = AppRepoMapping.setup_logger(log_file)
        summary_log_file = os.path.join(logs_dir, f"AppRepoMappingSummary_log_{current_datetime}.txt")
        summary_logger = AppRepoMapping.create_summary_logger(summary_log_file)
        output_csv_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Summary.csv")
        if not os.path.exists(App_Repo_Mapping):
            print("Application to repository mapping information is missing, please refer README.md to create the mapping spreadsheet.")
            return
        mapping_excel_path = os.path.join(config_dir, f"App-Repo-Mapping.xlsx")
        add_action_column(App_Repo_Mapping, output_csv_file_path)
        AppRepoMapping.create_application_folders(App_Repo_Mapping, unzip_dir, src_dir_analyze, logger, summary_logger)
        # update_rescan_column(mapping_excel_path,output_csv_file_path,log_file)
    elif output_type==4:
        log_file = LoggerManager.get_logger("Fetch_New_Applications_", log_dir=logs_dir)
        app_log_file=LoggerManager.get_logger("create_new_Applications", log_dir=logs_dir)
        hl_applications_path=get_hl_applications_path(output_dir)
        output_csv_file_path = os.path.join(output_dir, f"{org_name}_New_applicatiosn.xlsx")
        find_new_applications(hl_applications_path,App_Repo_Mapping,log_file)
        create_applications_hl(highlight_base_url, app_log_file,App_Repo_Mapping,highlight_token,highlight_company_id)
    elif output_type == 9:  # New option for copying folder
        if not os.path.exists(csv_file_path):
            print(f"CSV file not found at {csv_file_path}. Please check your config.")
            return
        log_file_path = os.path.join(logs_dir, f"MainframeCopyAppendLog_{current_datetime}.log")
        mainframeCopyAppend_to_analyzed_from_csv(csv_file_path, log_file_path)
    elif output_type==5:
        hl_applications_path=get_hl_applications_path(output_dir)
        identify_to_be_deleted_apps(hl_applications_path,output_dir,logs_dir,App_Repo_Mapping)
    elif output_type == 13:
        parent_folder = src_dir_analyze
        if os.path.exists(parent_folder):
            long_paths = find_long_paths(parent_folder)
            if long_paths:
                # Create the output csv file
                long_path_csv_file = os.path.join(output_dir, f'Application_Long_Path_{current_datetime}.csv')
                with open(long_path_csv_file, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['Application Name', 'Long Path'])
                    for app_name, app_path in long_paths:
                        writer.writerow([app_name, app_path])
                print(f"Applications Long path saved to '{long_path_csv_file}'.\n")
            else:
                print("\n No paths exceeding the length limit found.\n")
        else:
            print("The specified parent folder does not exist.\n")
    
    elif output_type == 11:
        try:
            HLScanAndOnboard.main()
        except Exception as e:
            logging.error(f'{e}')
    
    elif output_type == 8:

        # 1. Download Metadata for GitHub organization
        output_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Metadata.json")
        log_file_path = os.path.join(logs_dir, f"{org_name}_Metadatadownload_{current_datetime}.log")
        output_csv_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Summary.csv")
        mapping_excel_path = os.path.join(config_dir, f"App-Repo-Mapping.xlsx")
        get_all_repo_metadata(org_name, token, output_file_path, log_file_path)
        json_to_csv(output_file_path, output_csv_file_path)
        modify_archive_urls(output_csv_file_path)
        add_new_columns_to_csv(output_csv_file_path)
        update_download_column(output_csv_file_path, last_refresh_date)
        add_action_column(mapping_excel_path, output_csv_file_path)
        print(f"Refer Log file {log_file_path} for download log and time to download Metadata.")
        print(f"CSV file generated {output_csv_file_path} with summary of repositories which can be used for downloading source code(Task-2).\n")

        # 2. Download source code for all repositories in the organization in batches
        output_csv_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Summary.csv")
        if not os.path.exists(output_csv_file_path):
            print("Please run option 1 to download metadata first.")
            return
        if not check_column_exists(output_csv_file_path, 'batch_number'):
            print(f"Column 'batch_number' does not exist in file {output_csv_file_path}. Create a new column with the name 'batch_number' and enter the number of the batch you want to run for download.")
            return
        if not check_column_exists(output_csv_file_path, 'repo_archive_download_api'):
            print(f"Column 'repo_archive_download_api' does not exist in file {output_csv_file_path}. Review the CSV file and rerun option 1.")
            return
        if not check_column_exists(output_csv_file_path, 'name'):
            print(f"Column 'name' does not exist in file {output_csv_file_path}. Review the CSV file and rerun option 1.")
            return
        output_csv_file_path = os.path.join(output_dir, f"{org_name}_Repositories_Summary.csv")
        data = read_csv_data(output_csv_file_path) 
        batches = []
        num_of_batches = int(data[-1][9])
        for i in range(num_of_batches):
            batch = []
            for repository in data:
                if int(repository[9]) == i+1:
                    batch.append(repository)
            batches.append(batch)
        # Process batches using multi-threading
        threads = []
        # Define the directory name
        directory = Path(output_dir+"\csv_for_each_batch")
        # Create the directory
        directory.mkdir(parents=True, exist_ok=True)
        print(f"Directory '{directory}' created successfully.")

        if os.path.exists(directory):
            # List all files in the directory
            files_in_directory = os.listdir(directory)
            # Loop through the files and delete each one
            for file in files_in_directory:
                file_path = os.path.join(directory, file)
                # Check if it's a file and delete it
                if os.path.isfile(file_path):
                    os.remove(file_path)
                    print(f"Deleted file: {file_path}")
        
        for i, batch in enumerate(batches, start=1):
            start_end_log_file = os.path.join(logs_dir, f"RepoDownloadTime_batch_{i}_{current_datetime}.txt")
            processing_log_file = os.path.join(logs_dir, f"RepoDownloadStatusLog_batch_{i}_{current_datetime}.txt")
            output_csv_file_path_batch = os.path.join(directory, f"{org_name}_Repositories_Summary_batch_{i}.csv")
            split_csv_based_on_batch_num(i, output_csv_file_path, output_csv_file_path_batch)

            if not os.path.exists(start_end_log_file):
                with open(start_end_log_file, "w") as start_end_log:
                    start_end_log.write("Start Time\tEnd Time\tTotal Time Taken\n")            
            if not os.path.exists(processing_log_file):
                with open(processing_log_file, "w") as processing_log:
                    processing_log.write("Timestamp\tMessage\n")
            open(start_end_log_file, 'w').close()
            open(processing_log_file, 'w').close()
            thread = threading.Thread(target=download_in_batch, args=(batch, i, src_dir, token, start_end_log_file, processing_log_file, output_csv_file_path, output_csv_file_path_batch))
            threads.append(thread)
        # Start threads
        for t in threads:
            t.start()
        # Join threads to the main thread
        for t in threads:
            t.join()

        combine_all_batch_csv_files(directory, output_csv_file_path)   

        # 3. Unzip the downloaded source code
        try:
            UnzipFile.unzip_code(src_dir, unzip_dir, os.path.join(logs_dir, f"Unzip_Execution_{current_datetime}.log"), os.path.join(logs_dir, f"Unzip_Time_{current_datetime}.log"))
        except Exception as e:
            print(f"Error occurred during extraction: {e}")

        # 4. Create application folders and move repositories
        log_file=os.path.join(logs_dir, f"AppRepoMapping_{current_datetime}.log")
        logger = AppRepoMapping.setup_logger(log_file)
        summary_log_file = os.path.join(logs_dir, f"AppRepoMappingSummary_log_{current_datetime}.txt")
        summary_logger = AppRepoMapping.create_summary_logger(summary_log_file)
        if not os.path.exists(App_Repo_Mapping):
            print("Application to repository mapping information is missing, please refer README.md to create the mapping spreadsheet.")
            return
        mapping_excel_path = os.path.join(config_dir, f"App-Repo-Mapping.xlsx")
        add_action_column(mapping_excel_path, output_csv_file_path)
        AppRepoMapping.create_application_folders(App_Repo_Mapping, unzip_dir, src_dir_analyze, logger, summary_logger)

        # 5. Copy or Append Mainframe folder to analyzed directory
        mainframeCopyAppend_to_analyzed_from_csv(csv_file_path, os.path.join(logs_dir, f"mainframe_copy_append_log_{current_datetime}.log"))

        # 6. Get applications long path
        parent_folder = src_dir_analyze
        if os.path.exists(parent_folder):
            long_paths = find_long_paths(parent_folder)
            if long_paths:
                # Create the output csv file
                long_path_csv_file = os.path.join(output_dir, f'Application_Long_Path_{current_datetime}.csv')
                with open(long_path_csv_file, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['Application Name', 'Long Path'])
                    for app_name, app_path in long_paths:
                        writer.writerow([app_name, app_path])
                print(f"Applications Long path saved to '{long_path_csv_file}'.\n")
            else:
                print("\n No paths exceeding the length limit found.\n")
        else:
            print("The specified parent folder does not exist.\n")

        # 7. Trigger CAST Highlight onboarding for the source code
        try:
            HLScanAndOnboard.main()
        except Exception as e:
            logging.error(f'{e}')
    elif output_type == 10:
        # Step 1: Setup loggers
        rescan_logger = LoggerManager.get_logger("re_export_HL_existing_data_", log_dir=logs_dir)
        app_logger = LoggerManager.get_logger("app_txt", log_dir=logs_dir)

        # Step 2: Prepare API details
        url = f"{highlight_base_url}/WS2/domains/{highlight_company_id}/applications"
        headers = {"Authorization": f"Bearer {highlight_token}"}

        # Step 3: Prepare paths
        rescan_summary = os.path.join(output_dir, "Rescan_applications_summary.xlsx")

        # Step 4: Find the latest LM_Highlight_AppDetails_ file
        latest_hl_data = None
        for file in os.listdir(output_dir):
            if file.startswith("LM_Highlight_AppDetails_") and file.endswith(".xlsx"):
                latest_hl_data = os.path.join(output_dir, file)
                break

        if not latest_hl_data:
            rescan_logger.error("No LM_Highlight_AppDetails_ file found in the output directory.")
            print("Error: LM_Highlight_AppDetails_ file not found.")
        else:
            from datetime import datetime
            current_date = datetime.now().strftime("%Y%m%d")
            fetch_and_save_applications(url, headers, output_dir, rescan_logger, current_date)

            # Step 6: Create Application TXT file for rescan apps
            create_app_txt(
                latest_hl_data=latest_hl_data,
                rescan_summary=rescan_summary,
                app_logger=app_logger,
                txt_output_path=highlight_application_mapping
            )

    else:
        print("Invalid choice.")

if __name__ == "__main__":
    try:    
        main()
    except Exception as e:
        print(f'{e}')
